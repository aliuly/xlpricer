#!/usr/bin/env python3
'''My Excel helper functions
'''

try:
  from icecream import ic
except ImportError:  # Graceful fallback if IceCream isn't installed.
  ic = lambda *a: None if not a else (a[0] if len(a) == 1 else a)  # noqa

import datetime
import openpyxl
import os
import sys

class XlUtils():
  '''Basic Excel class utilities'''
  def __init__(self, xlfile:str, rdonly=False, data=False):
    '''Constructor
    :param xlfile: Path to excel file to create
    '''
    self.vlists = dict()
    self.refs = dict()
    self.xlfile = xlfile

    if os.path.isfile(xlfile):
      self.xl = openpyxl.load_workbook(xlfile,read_only = rdonly, data_only = data)
      if rdonly: self.xlfile = None
      self.first_ws = False
    else:
      self.xl = openpyxl.Workbook()
      self.first_ws = True

  def close(self, savefile = None) -> None:
    '''Finalize writing to the created XLSX file'''
    if savefile is None:
      if self.xlfile is None: return
      savefile = self.xlfile
    sys.stderr.write(f'Writing {savefile} file..')
    self.xl.save(savefile)
    sys.stderr.write('.OK\n')

    # ~ while True:
      # ~ try:
        # ~ sys.stderr.write('Writing XLSX file..')
        # ~ self.xl.save(self.xlfile)
        # ~ sys.stderr.write('.OK\n')
      # ~ except xlsxwriter.exceptions.FileCreateError as e:
        # ~ decision = input("Exception caught in workbook.close(): %s\n"
                         # ~ "Please close the file if it is open in Excel.\n"
                         # ~ "Try to write file again? [Y/n]: " % e)
        # ~ if decision != 'n': continue
      # ~ break

  def add_vlist(self, name:str) -> None:
    '''Create a new validation list
    :param name: vlist to create
    '''
    if name in self.vlists: raise KeyError(f'Duplicate vlist {name}')
    self.vlists[name] = list()

  def vlist(self, name:str, value:str|None = None) -> None|list:
    '''Access to validation list

    :param name: vlist to access
    :param value: If specified, value is added to the list.  Otherwise the list is returned
    :returns: if no value was specified, returns the validation list
    '''
    if value is None: return self.vlists[name]
    if not value in self.vlists[name]:
      self.vlists[name].append(value)
      self.vlists[name].sort()

  def ref(self, *args, **kwargs) -> dict|str:
    '''Access to references

    If positional parameters are specified, it will return the references
    specified there.
    If kwargs are spcified, these will be used to define references.
    If nothing is specified, returns the refs dictionary
    '''
    if len(args) == 0 and len(kwargs) == 0: return self.refs
    for k,v in kwargs.items():
      self.refs[k] = v
    if len(args) == 1: return self.refs[args[0]]
    res = list()
    for i in args: res.append(self.refs[i])
    return res

  def rowrefs(self, r:int)->None:
    '''Update row references
    :param r: current row
    '''
    for f in list(self.refs.keys()):
      if not f.startswith('f_'): continue
      self.refs['#'+f] = self.refs[f] + str(r)
    self.refs['#'] = str(r)

  def ws(self, name:str) -> openpyxl.worksheet.worksheet.Worksheet:
    '''Accesor for the get_worksheet_by_name() function

    :param name: name of worksheet to get
    :returns: the given worksheept
    '''
    return self.xl[name]
  def add_worksheet(self, name:str) -> openpyxl.worksheet.worksheet.Worksheet:
    '''Accesor for the add_worksheet() function

    :param name: name of worksheet to get
    :returns: the given worksheept
    '''
    if self.first_ws:
      self.first_ws = False
      ws = self.xl.active
      ws.title = name
      return ws
    return self.xl.create_sheet(title=name)

  def load_fmt(self, ref:type, prefix = 'f_') -> None:
    '''Loads the defined formats into the open XLSX objec

    :param ref: class containing formats
    '''
    for k,v in ref.__dict__.items():
      if not k.startswith(prefix) or not isinstance(v,dict): continue
      if k in self.xl.named_styles:
        setattr(ref,k,str(k))
        continue

      custom_style = openpyxl.styles.NamedStyle(name=k)
      alignment = v['alignment'].copy() if 'alignment' in v else dict()
      if 'font' in v:
        custom_style.font = openpyxl.styles.Font(**v['font'])
      if 'fill' in v:
        if isinstance(v['fill'],str):
          custom_style.fill = openpyxl.styles.PatternFill(
                  start_color =  v['fill'],
                  end_color = v['fill'],
                  fill_type = 'solid',
          )
        else:
          custom_style.fill = openpyxl.styles.PatternFill(**v['fill'])
      if 'text_wrap' in v: alignment['wrap_text'] = v['text_wrap']
      if 'align' in v: alignment['horizontal'] = v['align']
      if 'valign' in v: alignment['vertical'] = v['valign']
      if 'border' in v:
        border = v['border']
        if not isinstance(border,dict) or not (('left' in border) or ('right' in border) or ('top' in border) or ('bottom' in border)):
          border = { 'left': border, 'right': border, 'top': border, 'bottom': border }
        for kw in list(border.keys()):
          if isinstance(border[kw],dict):
            border[kw] = openpyxl.styles.Side(**border[kw])
          else:
            border[kw] = openpyxl.styles.Side(style=border[kw])
        custom_style.border = openpyxl.styles.Border(**border)
      if 'num_format' in v:
        custom_style.number_format = v['num_format']

      if len(alignment) > 0: custom_style.alignment = openpyxl.styles.Alignment(**alignment)

      self.xl.add_named_style(custom_style)
      setattr(ref,k,str(k))

def today(fmt:str = '%Y-%m-%d') -> str:
  ''' Return today's date as a string

  This function is here because I didn't know where else to put it...
  :param fmt: Date format to use
  :returns: a string with today's date formated as YYYY-MM-DD (or as specified by format)
  '''
  return datetime.datetime.today().strftime(fmt)

def pick_default(opts:list, pref:str) -> str:
  ''' Helper function, select a suitable default from a choice list
  :param opts: list with option strings
  :param pref: Preferred value
  :returns: default value from the list of options
  '''
  if len(opts) == 0: return ''
  if pref in opts: return pref
  return opts[0]

def nuke_ws(ws:openpyxl.worksheet.worksheet.Worksheet) -> None:
  '''Delete all cells in worksheet

  :param ws: worksheet
  '''
  ws.data_validations.dataValidation.clear()
  for row in ws.iter_rows():
    for cell in row:
      cell.value = None
      cell.font = openpyxl.styles.Font()
      cell.fill = openpyxl.styles.PatternFill()
      cell.border = openpyxl.styles.Border()
      cell.alignment = openpyxl.styles.Alignment()


def write(ws:openpyxl.worksheet.worksheet.Worksheet, r:int, c:int, text:str, style:str|None = None) -> None:
  '''Write a cell value and optionally style

  :param ws: worksheet
  :param r: row
  :param c: column
  :param text: value to set
  :param style: style to use or set to None
  '''
  cell = ws.cell(r,c)
  if isinstance(text,str) and text.startswith('=='):
    text = text[1:]
    p = rowcol_to_cell(r,c)
    ws[p] = openpyxl.worksheet.formula.ArrayFormula(f'{p}:{p}', text)
  else:
    cell.value = text
  if style is not None: cell.style = style

def set_column_width(ws:openpyxl.worksheet.worksheet.Worksheet, c:int, width:float|int) -> None:
  '''Set the width of a column

  :param ws: worksheet
  :param c: column
  :param width: size of column
  '''
  col = openpyxl.utils.get_column_letter(c)
  ws.column_dimensions[col].width = width

def col_to_name(c:int, absolute:bool = False) -> str:
  '''Convert column index to name

  :param c: column
  :param absolute: Make an absolute reference
  :returns: column name
  '''
  return ('$' if absolute else '')+openpyxl.utils.get_column_letter(c)

def rowcol_to_cell(r:int, c:int, absrow:bool = False, abscol:bool = False) -> str:
  '''Convert row, column indeces to coordinate

  :param r: row
  :param c: column
  :param absrow: Make an absolute row reference
  :param abscol: Make an absolute col reference
  :returns: coordinate
  '''
  return '{dcol}{col}{drow}{row}'.format(
              dcol = '$' if abscol else '',
              col = openpyxl.utils.get_column_letter(c),
              drow = '$' if absrow else '',
              row = r)

def cell_to_rowcol(pos:str) -> [int,int]:
  '''Convert a coordinate to row,col indeces

  :param pos: coordinate
  :returns: row,col
  '''
  col,row = openpyxl.utils.cell.coordinate_from_string(pos)
  return row, openpyxl.utils.column_index_from_string(col)

def escape_excel_formula(value:str, forced:bool=True) -> str:
    """
    Escapes double quotes and commas in a string for Excel formulas.
    Double quotes are escaped by doubling them.
    Values with commas or double quotes are wrapped in double quotes.

    :param value: value to escape
    :returns: string with the escaped value
    """
    if '"' in value:  # Escape double quotes by doubling them
      value = value.replace('"', '""')
    if forced or (',' in value) or ('"' in value):  # Wrap the entire value in double quotes if it contains commas or double quotes
      value = f'"{value}"'
    return value

def data_validation_list(ws:openpyxl.worksheet.worksheet.Worksheet, r:int, c:int, vlist:list, hide_dropdown:bool=False, allow_blank:bool=True) -> None:
  '''Add data validation to a cell based on a list

  :param ws: worksheet
  :param r: row
  :param c: column
  :param vlist: either a string containing a formula or a list of options
  :param hide_dropdown: Do not show dropdown menu
  :param allow_blank: Set the allow blanks flag
  '''

  # Sanitize the list...
  if isinstance(vlist,list):
    # ~ f1 = ','.join([ escape_excel_formula(value,True) for value in vlist])
    f1 = ','.join(vlist)
    f1 = f'"{f1}"'
  else:
    f1 = vlist
  validator = openpyxl.worksheet.datavalidation.DataValidation(type = 'list',
                                                              formula1 = f1,
                                                              showErrorMessage=True,
                                                              showInputMessage=True,
                                                              showDropDown=hide_dropdown,
                                                              allowBlank=allow_blank,
                                                              )
  ws.add_data_validation(validator)
  validator.add(ws.cell(r,c))

def freeze_panes(ws:openpyxl.worksheet.worksheet.Worksheet, r:int, c:int) -> None:
  '''Freeze panes
  :param ws: worksheet
  :param r: row
  :param c: column
  '''
  ws.freeze_panes = rowcol_to_cell(r,c)


def autofilter(ws:openpyxl.worksheet.worksheet.Worksheet, t:int, l:int, b:int, r:int) -> None:
  '''Create an auto filter
  :param ws: worksheet
  :param t: top row
  :param l: left column
  :param b: bottom row
  :param r: right column
  '''
  ws.auto_filter.ref = rowcol_to_cell(t,l)+':'+rowcol_to_cell(b,r)

def group_columns(ws:openpyxl.worksheet.worksheet.Worksheet, start:int, end:int, **kwargs) ->None:
  '''Group columns
  :param ws: worksheet
  :param start: Starting column index
  :param end: Ending column index
  :param **kwargs: Optional arguments

  Optional arguments:

  - hide : boolean, start hidden
  - level : integer, outline level
  '''
  opts = dict()
  if isinstance(start,int): start = col_to_name(start)
  if isinstance(end,int): end = col_to_name(end)
  for k,v in dict(hide='hidden', level='outline_level').items():
    if k in kwargs: opts[v] = kwargs[k]
  ws.column_dimensions.group(start, end, **opts)


def group_rows(ws:openpyxl.worksheet.worksheet.Worksheet, start:int, end:int, **kwargs) ->None:
  '''Group columns
  :param ws: worksheet
  :param start: Starting row index
  :param end: Ending row index
  :param **kwargs: Optional arguments

  Optional arguments:

  - hide : boolean, start hidden
  - level : integer, outline level
  '''
  opts = dict()
  for k,v in dict(hide='hidden', level='outline_level').items():
    if k in kwargs: opts[v] = kwargs[k]
  ws.row_dimensions.group(start, end, **opts)

