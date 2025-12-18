#!/usr/bin/env python3
'''Write to volumetrics tab
'''
try:
  from icecream import ic
except ImportError:  # Graceful fallback if IceCream isn't installed.
  ic = lambda *a: None if not a else (a[0] if len(a) == 1 else a)  # noqa

import openpyxl

from . import xlu
from .constants import K
from .xlfmt import XlFmt

def sheet(xl:xlu.XlUtils, apidat:dict) -> None:
  '''Write to assumptions tab
  :param xl: xl utility object
  '''

  ws = xl.ws(K.WS_VOLUMES)
  xl.ref(WS_COMPONENT = K.WS_COMPONENT)

  SPACER = { 'h': [ None, 3 ] }
  '''Spacer column definition'''

  COLUMNS = [
    SPACER,
    {
      'h': [ "[", 4, XlFmt.f_header, 'f_v_lvol', True ],
      'f': XlFmt.f_num_c,
    },
    {
      'h': [ "Componets Tab Volume", 14, XlFmt.f_header, 'f_v_volume', True ],
      'f': XlFmt.f_num_c,
    },
    {
      'h': [ "]", 4, XlFmt.f_header, 'f_v_rvol', True ],
      'f': XlFmt.f_num_c,
    },
    SPACER,
    {
      'h': [ 'Tiered Volume', 10, XlFmt.f_header, 'f_v_tiered_vol', True ],
      'f': XlFmt.f_num_c,
    },
    SPACER,
    {
      'h': [ "Notes", 12, XlFmt.f_header, 'f_v_nots', True ],
      'f': XlFmt.f_info,
    },
    {
      'h': [ K.CN_DESC, 42, XlFmt.f_header, 'f_v_desc', True ],
      'f': XlFmt.f_desc,
    },
    {
      'h': [ K.CN_REGION, 7, XlFmt.f_header, 'f_v_reg', True  ],
      'f': XlFmt.f_text,
      'c': '={WS_REGION}',
    },
    SPACER,
    {
      'h': [ 'Row Idx', 6.5, XlFmt.f_syshdr, 'f_v_sku' ],
      'f': XlFmt.f_num_c,
      'c': '==IF(OR({#f_v_desc}="",{#f_v_reg}=""),"",MATCH(1,({PRICES_DESCS}={#f_v_desc})*({PRICES_REGION}={#f_v_reg}),0))',
    },
    {
      'h': [K.CN_TIERED_PRICE, 10, XlFmt.f_syshdr, 'f_v_price' ],
      'f': XlFmt.f_euro_4d,
      'c': '=IF({#f_v_sku}="","",'
              'INDEX({PRICES_TABLE},{#f_v_sku},{cm_priceAmount})'
        ')'
    },
    {
      'h': ['Unit', 15, XlFmt.f_syshdr, 'f_v_unit' ],
      'f': XlFmt.f_desc,
      'c': '=IF({#f_v_sku}="","",'
              'INDEX({PRICES_TABLE},{#f_v_sku},{cm_unit})'
        ')'
    },
    SPACER,
    {
      'h': ['Sub-total', 15, XlFmt.f_refhdr, 'f_v_total' ],
      'f': XlFmt.f_euro,
      'c': '={#f_v_tiered_vol}*{#f_v_price}',
    },
  ]
  '''Column definitions'''
  coloffs = len(COLUMNS)+1

  r = 1
  xlu.write(ws,r,1, 'Volume Pricing', XlFmt.f_title)
  
  r += 2

  for c in range(1,coloffs):
    ws_header(xl, r, c, COLUMNS[c-1]['h'])

  tiers = list(apidat['tiers'].keys())
  tiers.sort(key = lambda i: apidat['tiers'][i][K.COL_XLTITLE])
 
  crs = {}
  for c in COLUMNS:
    if len(c['h'])< 4: continue
    _, crs[c['h'][3]] = xlu.cell_to_rowcol(xl.ref(c['h'][3])+'1')
    
  targets = []
  
  for ttariff in tiers:
    r += 1
    # ~ ws.set_row(r,None,None,{'level':1, 'hidden': 1})
    xl.rowrefs(r)
    for c in range(1,len(COLUMNS)+1):
      ws_cell(xl,r,c, COLUMNS[c-1],'')

    c = crs['f_v_volume']
    f = ('=SUMIFS({WS_COMPONENT}!{f_qxh}:{f_qxh},'
            '{WS_COMPONENT}!{f_desc}:{f_desc},"="&{#f_v_desc},'
            '{WS_COMPONENT}!{f_reg}:{f_reg},"="&{#f_v_reg}'
        ')')
    # ~ ic(f.format(**xl.ref()))
    ws_cell(xl,r,c, COLUMNS[c-1],f)
    c = crs['f_v_tiered_vol']
    ws_cell(xl,r,c, COLUMNS[c-1],'=SUM({#f_v_lvol}:{#f_v_rvol})')
    c = crs['f_v_desc']
    ws_cell(xl,r,c, COLUMNS[c-1],apidat['tiers'][ttariff][K.COL_XLTITLE])
    c = crs['f_v_reg']
    ws_cell(xl,r,c, COLUMNS[c-1],apidat['tiers'][ttariff]['region'])
    c = crs['f_v_sku']
    ws_cell(xl,r,c, COLUMNS[c-1])
    c = crs['f_v_unit']
    ws_cell(xl,r,c, COLUMNS[c-1], apidat['tiers'][ttariff]['unit'])
    

    c = crs['f_v_price']
    f = '=IF({#f_v_tiered_vol}<>0,{#f_v_total}/{#f_v_tiered_vol},"")'
    ws_cell(xl,r,c, COLUMNS[c-1], f)
    
    xl.ref(**{
      '#volcell': xl.ref('#f_v_tiered_vol'),
      '#vregion': xl.ref('#f_v_reg'),
    })

    targets.append([
      apidat['tiers'][ttariff][K.COL_XLTITLE],
      apidat['tiers'][ttariff]['region'],
      xl.ref('#f_v_price'),
    ])

    first = None
    last = None
    for tariff in apidat['tiers'][ttariff][K.COL_XLTARIFFS]:
      r += 1
      xl.rowrefs(r)
      for c in range(crs['f_v_tiered_vol'],len(COLUMNS)+1):
        ws_cell(xl,r,c, COLUMNS[c-1],'')
      
      last = r
      if first is None: first = r

      c = crs['f_v_desc']
      ws_cell(xl,r,c, COLUMNS[c-1],tariff[K.COL_XLTITLE])
      c = crs['f_v_reg']
      ws_cell(xl,r,c, COLUMNS[c-1],tariff['region'])
      c = crs['f_v_sku']
      ws_cell(xl,r,c, COLUMNS[c-1])

      xl.ref(Tmin = 0 if tariff['fromOn'] == 0 else tariff['fromOn']-1,
             Tmax = tariff['upTo'])
      f = 'IF({#volcell}>={Tmin},{#volcell}-{Tmin},0)'
      if tariff['upTo']: f = 'IF({#volcell}>{Tmax},{Tmax}-{Tmin},'+f+')'
      # ~ if not R['Tmax'] is None: f = 'IF({#volcell}>{Tmax},{Tmax}-{Tmin},'+f+')'
      c = crs['f_v_tiered_vol']
      ws_cell(xl,r,c, COLUMNS[c-1], '='+f)
      c = crs['f_v_reg']
      ws_cell(xl,r,c, COLUMNS[c-1], '={#vregion}')
      c = crs['f_v_price']
      # ~ ws_cell(xl,r,c, COLUMNS[c-1], tariff['priceAmount'])
      ws_cell(xl,r,c, COLUMNS[c-1])
      c = crs['f_v_total']
      ws_cell(xl,r,c, COLUMNS[c-1])
      
    r += 1
    c = crs['f_v_total']
    f = f'=SUM({{f_v_total}}{first}:{{f_v_total}}{last})'
    ws_cell(xl,first-1,c, COLUMNS[c-1], f)

  ws_update_prices(xl, targets)

      
def ws_update_prices(xl:xlu.XlUtils, targets:list):
  ic(targets)
  # Update pricing sheet with target formulas
  ws = xl.ws(K.WS_PRICES)
  last_sku = xl.ref('LAST_SKU')
  cm_region = xl.ref('cm_region')
  cm_price = xl.ref('cm_priceAmount')
  for desc,region,cn in targets:
    # ~ ic(desc,region,cn)
    for r in range(1, last_sku):
      if ws.cell(r,cm_region).value == region and ws.cell(r,1).value == desc:
        xlu.write(ws,r,cm_price, f'={K.WS_VOLUMES}!{cn}')
      # ~ if   xlu.write(ws,r,c, content, fmt)


def find_targets(xl:xlu.XlUtils) -> list:
  ws = xl.ws(K.WS_VOLUMES)
  
  cols = {
    K.CN_DESC: None,
    K.CN_REGION: None,
    K.CN_TIERED_PRICE: None,
  }
  targets = []
  
  for row in ws.iter_rows():
    if not all(value is not None for value in cols.values()):
      # Look for column names...
      c = 0
      for cell in row:
        if cell.value is not None and str(cell.value) in cols:
          cols[str(cell.value)] = c
        c += 1
      # ~ ic(cols)
    else:
      if not all(row[cc].value is not None for cc in cols.values()) or str(row[cols[K.CN_REGION]].value).startswith('='): continue

      targets.append([
        str(row[cols[K.CN_DESC]].value),
        str(row[cols[K.CN_REGION]].value),
        '$'+row[cols[K.CN_TIERED_PRICE]].coordinate,
      ])
  return targets


def ws_header(xl:xlu.XlUtils,r:int,c:int,hdef:list)->None:
  '''Write a header cell (formatting columns on the way)

  :param xl: xl utility object
  :param r: row
  :param c: column
  :param hdef: column header definition
  '''
  ws = xl.ws(K.WS_VOLUMES)

  # ~ ic(XlBomCols.COLUMNS[c])
  h = hdef[0] if len(hdef) > 0 else None
  w = hdef[1] if len(hdef) > 1 else None
  fmt = hdef[2] if len(hdef) > 2 else XlFmt.f_header
  fld_id = hdef[3] if len(hdef) > 3 else None
  hprot = hdef[4] if len(hdef) > 4 else False
  xlu.set_column_width(ws,c,w)

  if h is None: return
  xlu.write(ws,r,c, h, fmt)
  if not fld_id is None: xl.ref(**{fld_id: xlu.col_to_name(c,True)})
  if hprot: xlu.data_validation_list(ws, r, c, [ h ], True, False)

def ws_cell(xl:xlu.XlUtils,r:int,c:int, coldef:dict, o:str = None)->None:
  '''Write a cell

  Will write data on a cell, either from the given parameter `o` or use
  the default from the header definition.  Usually, the default is for
  creating cells with some default calculation.

  In either case, `o` can be:

  - `None`, will use the pre-defined default for the column.
  - _callable_: A lambda that takes a single parameter `ref`.
  - _str_ : A string, which if it starts with `=` it will be written
    as a formula, otherwise it is written as a value.  The string
    will be formated using `str.format` with the `ref` dict being
    passed as named arguments.

  :param xl: xl utility object
  :param r: row
  :param c: column
  :param coldef: Column defintion
  '''
  ws = xl.ws(K.WS_VOLUMES)

  content = (coldef['c'] if 'c' in coldef else None) if o is None else o
  fmt = coldef['f'] if ('f' in coldef) else XlFmt.f_def_data

  if callable(content):
    content = content(xl.ref())
  elif isinstance(content,str):
    content = content.format(**xl.ref())

  xlu.write(ws,r,c, content, fmt)

  if 'validate-list' in coldef:
    xlu.data_validation_list(ws, r, c, coldef['validate-list'])
